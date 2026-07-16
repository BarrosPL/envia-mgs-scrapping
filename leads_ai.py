from __future__ import annotations

import csv
import json
import os
import re
from dataclasses import dataclass
from pathlib import Path
from urllib.parse import urlparse

from dotenv import load_dotenv
from openai import OpenAI
from openpyxl import load_workbook


USERNAME_RE = re.compile(r"^[A-Za-z0-9._]{1,30}$")
URL_RE = re.compile(r"https?://\S+", re.IGNORECASE)


@dataclass(frozen=True)
class Lead:
    lead_id: str
    username: str
    qualification_reason: str
    intent: str
    confidence: float


class AIMessageError(RuntimeError):
    pass


def as_bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "sim", "yes", "s"}


def load_leads(
    path: Path,
    sheet_name: str = "Leads Instagram",
    min_confidence: float = 80,
) -> list[Lead]:
    if not path.exists():
        raise FileNotFoundError(f"planilha não encontrada: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"aba não encontrada: {sheet_name}")
        rows = workbook[sheet_name].iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            return []
        headers = [str(value).strip() if value is not None else "" for value in header_row]
        required = {
            "id",
            "instagram_username",
            "motivo_qualificacao",
            "intencao_detectada",
            "confianca_ia",
            "qualificado",
            "contato_realizado",
        }
        missing = required.difference(headers)
        if missing:
            raise ValueError(f"colunas ausentes na planilha: {', '.join(sorted(missing))}")

        leads: list[Lead] = []
        seen: set[str] = set()
        for row_number, values in enumerate(rows, 2):
            record = dict(zip(headers, values))
            if not as_bool(record.get("qualificado")) or as_bool(record.get("contato_realizado")):
                continue
            username = str(record.get("instagram_username") or "").strip().removeprefix("@").lower()
            if not USERNAME_RE.fullmatch(username):
                raise ValueError(f"username inválido na linha {row_number}: {username!r}")
            if username in seen:
                continue
            try:
                confidence = float(record.get("confianca_ia"))
            except (TypeError, ValueError) as exc:
                raise ValueError(f"confiança inválida na linha {row_number}") from exc
            reason = str(record.get("motivo_qualificacao") or "").strip()
            if confidence < min_confidence or not reason:
                continue
            seen.add(username)
            leads.append(
                Lead(
                    lead_id=str(record.get("id") or "").strip(),
                    username=username,
                    qualification_reason=reason,
                    intent=str(record.get("intencao_detectada") or "").strip(),
                    confidence=confidence,
                )
            )
        return leads
    finally:
        workbook.close()


def load_sent_usernames(path: Path) -> set[str]:
    if not path.exists():
        return set()
    with path.open(newline="", encoding="utf-8-sig") as file:
        return {
            str(row.get("username") or "").strip().lower().removeprefix("@")
            for row in csv.DictReader(file)
            if str(row.get("status") or "").strip().lower() == "enviado"
        }


def validate_service_url(value: str) -> str:
    value = value.strip()
    parsed = urlparse(value)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        raise ValueError("CUSTOMER_SERVICE_URL deve ser uma URL http(s) válida")
    return value


class OpenAIMessageGenerator:
    INSTRUCTIONS = """
Você redige mensagens iniciais de atendimento comercial em português do Brasil.
Crie somente o corpo da mensagem, em texto simples, com 2 frases curtas e tom de conversa.
Pode começar com uma saudação natural, como "Olá! Tudo bem?".
Use o motivo de qualificação apenas para descobrir o tema geral, como cidadania italiana,
cidadania portuguesa, visto ou documentação. Não repita a pergunta, condição familiar,
localização ou outro detalhe específico que originou a qualificação.
Diga de forma clara que a pessoa respondeu a uma pesquisa realizada por nossa equipe e, nessa
resposta, demonstrou interesse em conhecer nossos serviços. Em seguida, descreva de forma natural
qual interesse foi demonstrado, usando o motivo de qualificação como base, por exemplo: entender
as possibilidades de cidadania italiana, buscar orientação sobre documentação ou conhecer nosso
atendimento. Não copie o motivo literalmente nem mencione classificação, pontuação ou análise de IA.
Não use frases vagas como "chegamos ao seu perfil" ou "talvez esse tema faça sentido". Não cite
uma publicação ou comentário específico e não diga que monitoramos suas atividades.
Seja cordial, humano e não invasivo. Não invente nomes, fatos, documentos, preços, prazos,
elegibilidade, garantias ou aconselhamento jurídico. Não inclua links, markdown, assinatura,
username, pedido de dados pessoais ou CTA; o convite final será acrescentado pelo sistema.
""".strip()

    def __init__(self, api_key: str, service_url: str, model: str = "gpt-5.6-luna"):
        if not api_key.strip() or api_key.strip() == "cole_sua_chave_aqui":
            raise ValueError("OPENAI_API_KEY não configurada no .env")
        self.client = OpenAI(api_key=api_key.strip())
        self.service_url = validate_service_url(service_url)
        if urlparse(self.service_url).hostname in {"seu-link-de-atendimento.example"}:
            raise ValueError("CUSTOMER_SERVICE_URL ainda contém o valor de exemplo")
        self.model = model.strip() or "gpt-5.6-luna"

    @classmethod
    def from_env(cls) -> "OpenAIMessageGenerator":
        load_dotenv()
        return cls(
            api_key=os.getenv("OPENAI_API_KEY", ""),
            service_url=os.getenv("CUSTOMER_SERVICE_URL", ""),
            model=os.getenv("OPENAI_MODEL", "gpt-5.6-luna"),
        )

    def generate(self, lead: Lead) -> str:
        context = {
            "intencao_detectada": lead.intent,
            "motivo_qualificacao": lead.qualification_reason,
        }
        try:
            response = self.client.responses.create(
                model=self.model,
                instructions=self.INSTRUCTIONS,
                input=json.dumps(context, ensure_ascii=False),
                max_output_tokens=220,
            )
        except Exception as exc:
            raise AIMessageError(f"falha na OpenAI API: {type(exc).__name__}") from exc

        body = response.output_text.strip().strip('"')
        if not body:
            raise AIMessageError("a IA retornou uma mensagem vazia")
        if URL_RE.search(body):
            raise AIMessageError("a IA incluiu um link não autorizado")
        if len(body) > 700:
            raise AIMessageError("a IA retornou uma mensagem longa demais")

        cta = (
            "Se esse assunto ainda fizer sentido para você, podemos conversar com calma e ajudar "
            f"a entender melhor por onde começar. É só chamar nossa equipe por aqui: {self.service_url}"
        )
        return f"{body}\n\n{cta}"
